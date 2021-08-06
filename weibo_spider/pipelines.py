# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import json

from openpyxl import Workbook


class WeiboSpiderPipeline:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.create_sheet('weibo',0)
        self.ws_c = self.wb.create_sheet('comment',1)
        self.ws.append(['user_id', 'user_name', 'weibo_id', 'text', 'article_url', 'pic_urls', 'video_url',
                        'created_time', 'like_count', 'comment_count', 'repost_count', 'theme', '@user', 'weibo_url', 'tool',
                        'reposted_user_id', 'reposted_user_name', 'repost_text', 'repost_pic_urls', 'repost_article_url', 'repost_video_url', 'repost_created_time'
                        'repost_tool','repost_@user', 'repost_like_count', 'repost_comment_count', 'repost_retweet_count', 'repost_theme', 'repost_url'])
        self.ws_c.append(['weibo_id', 'comment', 'like_count'])

    def process_item(self, item, spider):
        line = [item.id, item.name, item.weibo_id, item.text, item.article_url, item.pic_url, item.video_url,
                item.post_time, item.attitudes_count, item.comments_count, item.repost_count, item.topics,
                item.at_user, item.weibo_url, item.post_tool, 
                item.repost['user_id'] if item.repost else '', item.repost['screen_name'] if item.repost else '',
                item.repost['text'] if item.repost else '', item.repost['pics'] if item.repost else '',
                item.repost['article_url'] if item.repost else '', item.repost['video_url'] if item.repost else '',
                item.repost['created_at'] if item.repost else '', item.repost['source'] if item.repost else '',
                item.repost['at_users'] if item.repost else '', item.repost['attitudes_count'] if item.repost else '',
                item.repost['comments_count'] if item.repost else '',
                item.repost['reposts_count'] if item.repost else '', item.repost['topics'] if item.repost else '',
                item.repost['url'] if item.repost else '']
        self.ws.append(line)
        if item.comments:
            comments_line = [item.weibo_id]
            for comment, like_count in item.comments:
                comments_line.extend([comment, like_count])
            self.ws_c.append(comments_line) 
        self.wb.save('result.xlsx')
        return item
